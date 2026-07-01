#!/usr/bin/env python3
"""
Build the canonical SAGP SQLite database from the frozen Master sheet.

App 2 in the SAGP workflow:
    Reconciliation workbook -> canonical SQLite database

The generated workbook remains an output of App 1.  This script treats the
Master sheet as the migration contract and creates the database that App 3
will manage going forward.
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import re
import sqlite3
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import load_workbook

WORKBOOK_FORMAT_VERSION = "SAGP_MASTER_V0_5"
DEFAULT_SHEET = "Master"

REQUIRED_COLUMNS = [
    "PersonID",
    "DisplayName",
    "FirstName",
    "MiddleName",
    "LastName",
    "Suffix",
    "Institution",
    "Title",
    "PrimaryEmail",
    "Phone",
    "City",
    "StateProvince",
    "PostalCode",
    "Country",
    "OriginalMembershipCode",
    "MembershipStatus",
    "LastPaidYear",
    "CodeHistory",
    "AppearsIn",
]

COLUMN_TO_DB = {
    "PersonID": "person_id",
    "DisplayName": "display_name",
    "FirstName": "first_name",
    "MiddleName": "middle_name",
    "LastName": "last_name",
    "Suffix": "suffix",
    "Institution": "institution",
    "Title": "title",
    "PrimaryEmail": "primary_email",
    "Phone": "phone",
    "City": "city",
    "StateProvince": "state_province",
    "PostalCode": "postal_code",
    "Country": "country",
    "OriginalMembershipCode": "original_membership_code",
    "MembershipStatus": "membership_status",
    "LastPaidYear": "last_paid_year",
}


def clean_cell(value) -> str | None:
    """Normalize spreadsheet cell values for SQLite storage."""
    if value is None:
        return None
    text = str(value).strip()
    if text == "":
        return None
    # Collapse accidental internal whitespace, but preserve meaningful punctuation.
    text = re.sub(r"\s+", " ", text)
    return text


def split_history(value) -> list[str]:
    """Split semicolon/comma-delimited workbook history fields."""
    text = clean_cell(value)
    if not text:
        return []
    parts = re.split(r"\s*;\s*|\s*,\s*", text)
    seen: set[str] = set()
    out: list[str] = []
    for part in parts:
        part = clean_cell(part)
        if part and part not in seen:
            seen.add(part)
            out.append(part)
    return out


def read_master_rows(workbook_path: Path, sheet_name: str = DEFAULT_SHEET) -> tuple[list[dict[str, str | None]], list[str]]:
    """Read rows from the frozen Master sheet and return normalized dictionaries."""
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Workbook {workbook_path} does not contain a sheet named {sheet_name!r}.")

    ws = wb[sheet_name]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = [clean_cell(v) for v in next(rows_iter)]
    except StopIteration as exc:
        raise ValueError(f"Sheet {sheet_name!r} is empty.") from exc

    missing = [c for c in REQUIRED_COLUMNS if c not in headers]
    if missing:
        raise ValueError(
            "Master sheet does not match the frozen App 2 contract. "
            f"Missing required columns: {missing}. Found: {headers}"
        )

    header_index = {h: i for i, h in enumerate(headers) if h}
    master_rows: list[dict[str, str | None]] = []
    warnings: list[str] = []
    seen_person_ids: set[str] = set()
    skipped_blank_rows = 0

    for excel_row_number, row in enumerate(rows_iter, start=2):
        values = {col: clean_cell(row[header_index[col]]) for col in REQUIRED_COLUMNS}
        if not any(values.values()):
            skipped_blank_rows += 1
            continue

        person_id = values.get("PersonID")
        display_name = values.get("DisplayName")
        if not person_id:
            warnings.append(f"Row {excel_row_number}: skipped row with no PersonID")
            continue
        if person_id in seen_person_ids:
            warnings.append(f"Row {excel_row_number}: duplicate PersonID {person_id}; later row skipped")
            continue
        if not display_name:
            warnings.append(f"Row {excel_row_number}: PersonID {person_id} has no DisplayName")
            values["DisplayName"] = person_id

        seen_person_ids.add(person_id)
        values["__excel_row_number"] = str(excel_row_number)
        master_rows.append(values)

    if skipped_blank_rows:
        warnings.append(f"Skipped {skipped_blank_rows} blank rows")
    return master_rows, warnings


def load_schema(conn: sqlite3.Connection, schema_path: Path) -> None:
    conn.executescript(schema_path.read_text(encoding="utf-8"))


def build_database(
    workbook_path: Path,
    db_path: Path,
    sheet_name: str = DEFAULT_SHEET,
    schema_path: Path | None = None,
    overwrite: bool = False,
) -> dict[str, object]:
    """Create a SQLite database from the Master sheet."""
    workbook_path = workbook_path.resolve()
    db_path = db_path.resolve()
    schema_path = schema_path or Path(__file__).resolve().parents[1] / "schema" / "sagp_members_schema.sql"

    if not workbook_path.exists():
        raise FileNotFoundError(workbook_path)
    if db_path.exists():
        if overwrite:
            db_path.unlink()
        else:
            raise FileExistsError(f"Database already exists: {db_path}. Use --overwrite to replace it.")

    rows, warnings = read_master_rows(workbook_path, sheet_name)
    now = dt.datetime.now(dt.timezone.utc).isoformat(timespec="seconds")

    db_path.parent.mkdir(parents=True, exist_ok=True)
    with sqlite3.connect(db_path) as conn:
        conn.execute("PRAGMA foreign_keys = ON")
        load_schema(conn, schema_path)

        member_records = []
        source_records = []
        code_records = []

        for row in rows:
            record = {db_col: row.get(xlsx_col) for xlsx_col, db_col in COLUMN_TO_DB.items()}
            if record.get("last_paid_year") is not None:
                try:
                    record["last_paid_year"] = int(record["last_paid_year"])
                except (TypeError, ValueError):
                    record["last_paid_year"] = None
            record["created_at"] = now
            record["updated_at"] = now
            record["active"] = 1
            record["notes"] = None
            member_records.append(record)

            person_id = row["PersonID"]
            for src in split_history(row.get("AppearsIn")):
                source_records.append((person_id, src))
            for code in split_history(row.get("CodeHistory")):
                code_records.append((person_id, code))

        conn.executemany(
            """
            INSERT INTO members (
                person_id, display_name, first_name, middle_name, last_name, suffix,
                institution, title, primary_email, phone, city, state_province,
                postal_code, country, membership_status, original_membership_code,
                last_paid_year, active, notes,
                created_at, updated_at
            ) VALUES (
                :person_id, :display_name, :first_name, :middle_name, :last_name, :suffix,
                :institution, :title, :primary_email, :phone, :city, :state_province,
                :postal_code, :country, :membership_status, :original_membership_code,
                :last_paid_year, :active, :notes,
                :created_at, :updated_at
            )
            """,
            member_records,
        )
        conn.executemany(
            "INSERT INTO source_appearances (person_id, source_file) VALUES (?, ?)",
            source_records,
        )
        conn.executemany(
            "INSERT INTO membership_code_history (person_id, code) VALUES (?, ?)",
            code_records,
        )
        conn.executemany(
            "INSERT OR REPLACE INTO schema_info (key, value) VALUES (?, ?)",
            [
                ("app", "SAGP App 2 Database Builder"),
                ("workbook_format_version", WORKBOOK_FORMAT_VERSION),
                ("created_at", now),
                ("source_workbook", str(workbook_path)),
                ("source_sheet", sheet_name),
            ],
        )
        conn.execute(
            """
            INSERT INTO import_log (
                imported_at, input_workbook, input_sheet, workbook_format_version,
                row_count, member_count, skipped_blank_rows, warnings
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                now,
                str(workbook_path),
                sheet_name,
                WORKBOOK_FORMAT_VERSION,
                len(rows),
                len(member_records),
                sum(1 for w in warnings if "blank rows" in w.lower()),
                json.dumps(warnings, indent=2),
            ),
        )
        conn.commit()

        counts = {
            "members": conn.execute("SELECT COUNT(*) FROM members").fetchone()[0],
            "source_appearances": conn.execute("SELECT COUNT(*) FROM source_appearances").fetchone()[0],
            "membership_code_history": conn.execute("SELECT COUNT(*) FROM membership_code_history").fetchone()[0],
        }

    return {
        "database": str(db_path),
        "input_workbook": str(workbook_path),
        "input_sheet": sheet_name,
        "workbook_format_version": WORKBOOK_FORMAT_VERSION,
        "counts": counts,
        "warnings": warnings,
    }


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build the SAGP SQLite database from the reconciled Master sheet.")
    parser.add_argument("--input", "-i", required=True, type=Path, help="Path to SAGP_Reconciliation.xlsx")
    parser.add_argument("--output", "-o", default=Path("output/sagp_members.db"), type=Path, help="Output SQLite database path")
    parser.add_argument("--sheet", default=DEFAULT_SHEET, help="Workbook sheet to read; default: Master")
    parser.add_argument("--overwrite", action="store_true", help="Replace the output database if it already exists")
    parser.add_argument("--json", action="store_true", help="Print machine-readable JSON summary")
    return parser.parse_args(argv)


def main(argv: Sequence[str] | None = None) -> int:
    args = parse_args(argv)
    result = build_database(args.input, args.output, args.sheet, overwrite=args.overwrite)
    if args.json:
        print(json.dumps(result, indent=2))
    else:
        print("SAGP database built successfully")
        print(f"  input:  {result['input_workbook']}")
        print(f"  sheet:  {result['input_sheet']}")
        print(f"  output: {result['database']}")
        for key, value in result["counts"].items():
            print(f"  {key}: {value}")
        warnings = result.get("warnings") or []
        if warnings:
            print("  warnings:")
            for warning in warnings[:10]:
                print(f"    - {warning}")
            if len(warnings) > 10:
                print(f"    ... {len(warnings) - 10} more")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
