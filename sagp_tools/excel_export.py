"""Excel workbook export for the SAGP reconciliation pipeline.

This module intentionally sits at the end of the pipeline. It does not parse,
merge, or reinterpret data; it only turns the already-generated rows into a
reviewable workbook for Betsy and future SAGP officers.

v0.3 note:
The first Excel exporter used openpyxl structured Table objects. Those are
nice, but Excel is much less forgiving of table XML than openpyxl/LibreOffice.
For messy contact exports, a safer workbook is produced by using ordinary
ranges with header styling, freeze panes, and autofilters. This version also
sanitizes all text values before writing cells.
"""

from __future__ import annotations

import math
import re
from pathlib import Path
from typing import Mapping, Sequence

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError as exc:  # pragma: no cover - friendly runtime message
    raise SystemExit(
        "Excel export requires openpyxl. Install it with: pip install openpyxl"
    ) from exc

Row = Mapping[str, object]

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(size=14, bold=True)
WRAP_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
CENTER_ALIGNMENT = Alignment(vertical="top", horizontal="center")

# Excel's worksheet XML cannot contain most ASCII control characters.
# Keep tab/newline/carriage return; remove the rest.
ILLEGAL_XML_RE = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")
EXCEL_MAX_TEXT = 32767

TEXT_HEAVY_FIELDS = {
    "Notes",
    "Address",
    "AppearsIn",
    "CodeHistory",
    "MergeReason",
    "SuggestedReason",
}
SHORT_FIELDS = {
    "PersonID",
    "RawRecordID",
    "SourceRow",
    "SourceRegion",
    "Suffix",
    "OriginalMembershipCode",
    "MergedRecordCount",
    "MergeConfidence",
    "SuggestedConfidence",
    "ReviewGroup",
    "CurrentPersonID",
}


def _safe_sheet_name(name: str) -> str:
    """Return an Excel-safe sheet name."""
    bad = set(r"[]:*?/\\")
    cleaned = "".join("_" if c in bad else c for c in str(name)).strip()
    return (cleaned or "Sheet")[:31]


def _safe_cell_value(value: object) -> object:
    """Return a value safe for writing into an xlsx cell.

    This prevents Excel repair warnings caused by invalid XML control
    characters, NaN/Inf values, and overlong cell text.
    """
    if value is None:
        return ""

    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return ""
        return value

    if isinstance(value, (int, bool)):
        return value

    text = str(value)
    text = ILLEGAL_XML_RE.sub(" ", text)

    # Excel interprets leading equals as formulas. These fields are data, so
    # protect literal text that begins with formula-like characters.
    if text.startswith(("=", "+", "-", "@")):
        text = "'" + text

    if len(text) > EXCEL_MAX_TEXT:
        text = text[: EXCEL_MAX_TEXT - 20] + " [TRUNCATED FOR EXCEL]"

    return text


def _write_table(ws, rows: Sequence[Row], fields: Sequence[str]) -> None:
    """Write rows into a worksheet and apply safe review formatting."""
    safe_fields = [_safe_cell_value(field) for field in fields]
    ws.append(safe_fields)

    for row in rows:
        ws.append([_safe_cell_value(row.get(field, "")) for field in fields])

    # Header styling.
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGNMENT

    ws.freeze_panes = "A2"
    if fields:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(fields))}{max(len(rows) + 1, 1)}"

    # Column widths. Avoid enormous notes/address columns.
    for idx, field in enumerate(fields, start=1):
        letter = get_column_letter(idx)
        sample_values = [str(field)] + [str(row.get(field, "")) for row in rows[:250]]
        sample_values = [ILLEGAL_XML_RE.sub(" ", value) for value in sample_values]
        max_len = max((len(v) for v in sample_values), default=len(str(field)))
        if field in TEXT_HEAVY_FIELDS:
            width = min(max(max_len + 2, 16), 48)
        elif field in SHORT_FIELDS:
            width = min(max(max_len + 2, 10), 18)
        else:
            width = min(max(max_len + 2, 12), 32)
        ws.column_dimensions[letter].width = width

    # Keep rows readable but not huge.
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = WRAP_ALIGNMENT


def _write_report_sheet(ws, report_text: str) -> None:
    ws["A1"] = "SAGP Reconciliation Report"
    ws["A1"].font = TITLE_FONT
    ws["A3"] = _safe_cell_value(report_text)
    ws["A3"].alignment = Alignment(vertical="top", wrap_text=True)
    ws.column_dimensions["A"].width = 90
    ws.row_dimensions[3].height = 280


def write_workbook(
    output_path: str | Path,
    *,
    master_rows: Sequence[Row],
    normalized_rows: Sequence[Row],
    duplicate_review_rows: Sequence[Row],
    merge_log_rows: Sequence[Row],
    source_summary_rows: Sequence[Row],
    code_summary_rows: Sequence[Row],
    report_text: str,
    master_fields: Sequence[str],
    normalized_fields: Sequence[str],
    review_fields: Sequence[str],
    merge_fields: Sequence[str],
    excluded_contact_rows: Sequence[Row] | None = None,
    excluded_contact_fields: Sequence[str] | None = None,
) -> Path:
    """Create the multi-sheet SAGP reconciliation workbook."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    _write_report_sheet(ws, report_text)

    sheets = [
        ("Master", master_rows, master_fields),
        ("DuplicateReview", duplicate_review_rows, review_fields),
        ("MergeLog", merge_log_rows, merge_fields),
        ("SourceSummary", source_summary_rows, ["SourceFile", "SourceRegion", "ImportedRows"]),
        ("CodeSummary", code_summary_rows, ["OriginalMembershipCode", "RowCount"]),
        ("RawNormalized", normalized_rows, normalized_fields),
    ]

    if excluded_contact_rows is not None and excluded_contact_fields is not None:
        sheets.append(("ExcludedContacts", excluded_contact_rows, excluded_contact_fields))

    for sheet_name, rows, fields in sheets:
        ws = wb.create_sheet(_safe_sheet_name(sheet_name))
        _write_table(ws, rows, fields)

    wb.save(output_path)
    return output_path
